#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Description: ONES工单夜间自动处理工具
@Version: 1.0.0
@Date: 2026-03-09
@Author: Lango AI Team

功能：
  - 读取ONES工单号（Excel / 命令行 / 交互式输入）
  - 调用 Claude Code CLI 执行工单任务
  - 结果回写Excel + 详细过程日志
  - 使用记录追踪上报

使用方式：
  # 命令行直接输入工单号
  python ones_task_runner.py --tickets ONES-12345 ONES-12346

  # 从Excel读取
  python ones_task_runner.py --excel tasks.xlsx

  # 交互式输入
  python ones_task_runner.py
"""

import argparse
import io
import json
import warnings

# 只抑制第三方库依赖版本警告（urllib3/chardet），不影响其他有价值的警告
warnings.filterwarnings("ignore", message=".*urllib3.*")
warnings.filterwarnings("ignore", message=".*chardet.*")
warnings.filterwarnings("ignore", category=DeprecationWarning, module="requests")
import logging
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from typing import Any, Dict, List

import requests
import yaml

# ============================================================
#  编码修复（Windows UTF-8）
# ============================================================
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# ============================================================
#  日志配置
# ============================================================
logger = logging.getLogger("ones_task_runner")
logger.setLevel(logging.INFO)

if sys.platform == 'win32':
    _console_handler = logging.StreamHandler(sys.stdout)
    _console_handler.setStream(io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8'))
else:
    _console_handler = logging.StreamHandler()

_formatter = logging.Formatter(
    '%(asctime)s  %(message)s',
    datefmt='%H:%M:%S'
)
_console_handler.setFormatter(_formatter)
logger.addHandler(_console_handler)


def ui(msg: str = ""):
    """纯 UI 输出（不带时间戳），用于 banner、分隔线等装饰内容"""
    print(msg, flush=True)


def display_width(text: str) -> int:
    """计算字符串在终端的显示宽度（中文/全角占2列，ASCII占1列）"""
    w = 0
    for ch in text:
        if ord(ch) > 0x7F:
            w += 2  # 中文、全角符号等
        else:
            w += 1
    return w


def header_line(title: str = "", width: int = 56):
    """输出带标题的分隔线（正确处理中文宽度）"""
    if title:
        title_w = display_width(title)
        pad = max(width - title_w - 4, 4)
        left = pad // 2
        right = pad - left
        ui(f"{'─' * left}  {title}  {'─' * right}")
    else:
        ui('─' * width)


# ============================================================
#  配置管理
# ============================================================
class Config:
    """配置管理器"""

    def __init__(self, config_path: str = "config.yaml"):
        self.config_path = config_path
        self._data = {}
        self.load()

    def load(self):
        """加载配置文件"""
        # 查找配置文件：优先当前目录，其次脚本所在目录
        paths_to_try = [
            self.config_path,
            os.path.join(os.path.dirname(os.path.abspath(__file__)), self.config_path),
        ]

        for path in paths_to_try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    self._data = yaml.safe_load(f) or {}
                logger.debug(f"配置已加载: {os.path.abspath(path)}")
                return

        logger.warning(f"未找到配置文件 {self.config_path}，使用默认配置")
        self._data = {}

    def get(self, key_path: str, default=None):
        """获取嵌套配置值，用点号分隔路径"""
        keys = key_path.split('.')
        value = self._data
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key)
            else:
                return default
            if value is None:
                return default
        return value


def normalize_ticket_id(raw: str) -> str:
    """标准化工单号：纯数字自动补 ONES- 前缀"""
    raw = raw.strip()
    if not raw:
        return raw
    # 已有 ONES- 前缀 → 统一大写
    if raw.upper().startswith('ONES-'):
        return 'ONES-' + raw[5:]
    # 纯数字 → 补 ONES-
    if raw.isdigit():
        return f'ONES-{raw}'
    return raw


def normalize_ticket_list(raw_list: list) -> list:
    """标准化工单号列表，支持逗号分隔的单字符串拆分"""
    result = []
    for item in raw_list:
        # 拆分逗号、分号、空格
        parts = [t.strip() for t in re.split(r'[,;\s]+', item) if t.strip()]
        for p in parts:
            result.append(normalize_ticket_id(p))
    return result



# ============================================================
#  Claude 执行引擎
# ============================================================
class ClaudeExecutor:
    """Claude Code 执行引擎（直接调用 claude -p）"""

    def __init__(self, config: Config):
        self.claude_path = config.get('claude.path', '') or 'claude'
        self.timeout = config.get('claude.timeout', 3600)
        self.max_turns = config.get('claude.max_turns', 100)  # [15.4] 增加轮数以支持多 subagent 调用

    def execute(self, task: Dict[str, Any]) -> Dict[str, Any]:
        """执行工单任务

        注意：不使用 /execute-task slash command，而是将 skill 模板内容
        直接内联到 prompt 中。这是因为 --permission-mode bypassPermissions
        与 slash command 组合会导致 claude CLI 死锁（已验证的 bug）。
        """
        task_id = task.get('taskId', '')
        params_json = self._build_task_params(task)

        # 加载 skill 模板并替换 $ARGUMENTS 占位符
        prompt = self._load_skill_template(params_json)

        args = [
            self.claude_path,
            '--permission-mode', 'bypassPermissions',
            '--max-turns', str(self.max_turns),
            '--verbose',
            '--output-format', 'stream-json',
            '-p', prompt,
        ]

        if sys.platform == 'win32':
            command = subprocess.list2cmdline(args)
        else:
            import shlex
            command = shlex.join(args)

        logger.debug(f"执行命令: {command[:200]}...")
        logger.info(f"⏳ 正在执行 {task_id}，预计数分钟...")

        # [15.5] 输出阶段标记（后端 task_executor 解析这些行推进时间线）
        print(f"[PHASE] agent_analyzing 开始分析工单 {task_id}", flush=True)

        start_time = time.time()

        try:
            # 流式执行：实时输出 Claude 的每一行分析过程
            # 使用 --output-format stream-json 让 claude 实时输出 JSON 事件
            # 解析事件并输出人类可读的进度信息
            import threading

            stdout_lines = []
            stderr_lines = []
            final_text_parts = []  # 收集最终文本（用于结果提取）

            proc = subprocess.Popen(
                command,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                universal_newlines=False,
            )

            def _drain_stderr():
                """后台线程：实时读取 stderr"""
                while True:
                    raw_line = proc.stderr.readline()
                    if not raw_line:
                        break
                    try:
                        line = raw_line.decode('utf-8', errors='replace').rstrip('\n\r')
                    except Exception:
                        line = str(raw_line)
                    if line:
                        stderr_lines.append(line)

            stderr_thread = threading.Thread(target=_drain_stderr, daemon=True)
            stderr_thread.start()

            # 主线程：逐行读取 stdout (readline 无缓冲延迟)
            while True:
                raw_line = proc.stdout.readline()
                if not raw_line:
                    break
                try:
                    line = raw_line.decode('utf-8', errors='replace').rstrip('\n\r')
                except Exception:
                    line = str(raw_line)
                if not line:
                    continue

                stdout_lines.append(line)

                # 解析 stream-json 事件并输出人类可读信息
                display_line = self._parse_stream_event(line, final_text_parts)
                if display_line:
                    print(display_line, flush=True)

            proc.wait(timeout=self.timeout)
            stderr_thread.join(timeout=5)

            duration = time.time() - start_time
            success = proc.returncode == 0
            if success:
                logger.info(f"✅ 任务成功: {task_id} (耗时 {duration:.1f}s)")
            else:
                logger.error(f"❌ 任务失败: {task_id} (退出码 {proc.returncode})")

            # 将收集的最终文本作为 stdout（兼容旧的结果提取逻辑）
            combined_text = ''.join(final_text_parts) if final_text_parts else '\n'.join(stdout_lines)

            return {
                'task_id': task_id,
                'success': success,
                'stdout': combined_text,
                'stderr': '\n'.join(stderr_lines),
                'returncode': proc.returncode,
                'duration': duration,
            }

        except subprocess.TimeoutExpired:
            proc.kill()
            duration = time.time() - start_time
            logger.error(f"⏰ 任务超时: {task_id} ({self.timeout}s)")
            return {
                'task_id': task_id,
                'success': False,
                'error': f'任务执行超时 ({self.timeout}s)',
                'duration': duration,
            }
        except Exception as e:
            duration = time.time() - start_time
            logger.error(f"💥 任务异常: {task_id} - {e}")
            return {
                'task_id': task_id,
                'success': False,
                'error': str(e),
                'exception_type': type(e).__name__,
                'duration': duration,
            }

    def _parse_stream_event(self, line: str, text_parts: list) -> str:
        """解析 claude --output-format stream-json 的事件行

        实际 stream-json 事件格式 (claude 2.1.x):
          {"type":"system","subtype":"init",...}
          {"type":"assistant","message":{"content":[{"type":"thinking","thinking":"..."}]}}
          {"type":"assistant","message":{"content":[{"type":"text","text":"..."}]}}
          {"type":"assistant","message":{"content":[{"type":"tool_use","name":"Bash","input":{}}]}}
          {"type":"tool_result","content":[{"type":"text","text":"..."}],...}
          {"type":"result","result":"...","total_cost_usd":0.05,...}
        """
        try:
            evt = json.loads(line)
        except (json.JSONDecodeError, ValueError):
            return line

        evt_type = evt.get('type', '')

        if evt_type == 'system':
            model = evt.get('model', '')
            return f"🚀 [初始化] 模型: {model}" if model else ""

        elif evt_type == 'assistant':
            message = evt.get('message', {})
            content_list = message.get('content', [])
            results = []
            for item in content_list:
                item_type = item.get('type', '')
                if item_type == 'thinking':
                    thinking = item.get('thinking', '')
                    if thinking:
                        snippet = thinking.replace('\n', ' ')[:200]
                        results.append(f"💭 [思考] {snippet}{'...' if len(thinking) > 200 else ''}")
                elif item_type == 'text':
                    text = item.get('text', '')
                    if text:
                        text_parts.append(text)
                        snippet = text.replace('\n', ' ')[:300]
                        results.append(f"📝 {snippet}")
                elif item_type == 'tool_use':
                    tool_name = item.get('name', 'unknown')
                    tool_input = item.get('input', {})
                    results.append(self._format_tool_use(tool_name, tool_input))
            return '\n'.join(results) if results else ""

        elif evt_type == 'tool_result':
            content_list = evt.get('content', [])
            for item in content_list:
                text = item.get('text', '') if isinstance(item, dict) else str(item)
                if text:
                    snippet = text.replace('\n', ' ')[:150]
                    return f"   ↪ 结果: {snippet}{'...' if len(text) > 150 else ''}"
            return ""

        elif evt_type == 'result':
            result_text = evt.get('result', '')
            if result_text:
                text_parts.append(result_text)
            cost = evt.get('total_cost_usd', 0)
            duration = evt.get('duration_ms', 0)
            turns = evt.get('num_turns', 0)
            return f"🏁 [完成] 耗时: {duration/1000:.1f}s, 轮次: {turns}, 消耗: ${cost:.4f}"

        elif evt_type == 'error':
            return f"❌ [错误] {evt.get('message', evt.get('error', str(evt)[:200]))}"

        return ""

    def _format_tool_use(self, tool_name: str, tool_input: dict) -> str:
        """格式化工具调用为人类可读文本"""
        n = tool_name.lower()
        if n == 'bash':
            return f"⚡ [执行命令] {tool_input.get('command', '')[:150]}"
        elif n in ('read', 'view'):
            return f"📖 [读取文件] {tool_input.get('file_path', tool_input.get('path', ''))}"
        elif n in ('write', 'edit'):
            return f"✏️  [写入文件] {tool_input.get('file_path', tool_input.get('path', ''))}"
        elif n in ('grep', 'search'):
            p = tool_input.get('pattern', tool_input.get('query', ''))
            return f"🔍 [搜索] {p}"
        elif n in ('glob', 'ls', 'list_dir'):
            return f"📂 [浏览] {tool_input.get('pattern', tool_input.get('path', ''))}"
        elif n.startswith('mcp__'):
            return f"🔧 [MCP: {tool_name.replace('mcp__', '')}] {str(tool_input)[:120]}"
        else:
            return f"🔧 [{tool_name}] {str(tool_input)[:150]}"

    def _load_skill_template(self, params_json: str) -> str:
        """加载 execute-task.md 模板并替换 $ARGUMENTS 占位符

        查找顺序：
        1. ~/.claude/commands/execute-task.md
        2. /opt/lango/commands/execute-task.md
        3. 脚本同目录下的 commands/execute-task.md
        4. 兜底：使用内置默认模板
        """
        search_paths = [
            os.path.join(os.path.expanduser("~"), ".claude", "commands", "execute-task.md"),
            "/opt/lango/commands/execute-task.md",
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "commands", "execute-task.md"),
        ]

        template = None
        for path in search_paths:
            if os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        template = f.read()
                    logger.debug(f"已加载 skill 模板: {path}")
                    break
                except Exception as e:
                    logger.debug(f"读取 {path} 失败: {e}")

        if template is None:
            # 兜底默认模板（含 subagent 管线指令）[15.3]
            logger.warning("未找到 execute-task.md 模板，使用内置默认模板（含 subagent 管线）")
            template = (
                "你是一个资深全栈软件工程师，正在自动执行一个ONES工单任务。"
                "以下是任务JSON：\n\n```json\n$ARGUMENTS\n```\n\n"
                "请按照以下 subagent 管线顺序完成任务：\n"
                "1. 使用 analyzer subagent 分析工单内容，理解问题根因，输出分析报告到 workspace/doc/{taskId}/analysis.md\n"
                "2. 使用 modifier subagent 根据分析报告修改代码，输出变更摘要到 workspace/doc/{taskId}/changes.md\n"
                "3. 使用 verifier subagent 验证代码修改结果，输出验证报告到 workspace/doc/{taskId}/verification.md\n"
                "4. 使用 reporter subagent 汇总生成最终报告到 workspace/doc/{taskId}/report/1.md\n\n"
                "每个阶段完成后请输出对应的阶段标记行（如: [PHASE] agent_analyzing_done 分析完成）。\n\n"
                "**重要**: 你已在 Agent-Teams 目录下启动，当前目录中的 .md 文件是项目规范和约束，"
                "请根据用户提示词和工单内容自行决定参考哪些规范文件。不要试图全部读取，按需查阅。\n\n"
                "开始执行任务。"
            )

        # 替换 $ARGUMENTS 占位符
        return template.replace("$ARGUMENTS", params_json)

    def _build_task_params(self, task: Dict[str, Any]) -> str:
        """构建任务JSON参数"""
        task_params = {
            "taskId": task.get('taskId', ''),
            "taskName": task.get('taskName', ''),
            "description": task.get('description', ''),
            "keywords": task.get('keywords', ''),
            "prompts": task.get('prompts', ''),
        }
        project_name = task.get('projectName', '')
        if project_name:
            task_params["projectName"] = project_name
        return json.dumps(task_params, ensure_ascii=False)


# ============================================================
#  Excel 管理器
# ============================================================
class ExcelManager:
    """Excel读写管理器"""

    # 列定义
    COL_TICKET = 1      # A列: 工单号
    COL_NOTE = 2        # B列: 补充说明
    COL_STATUS = 3      # C列: 状态
    COL_RESULT = 4      # D列: 执行结果
    COL_TIME = 5        # E列: 执行时间
    COL_PROCESS = 6     # F列: 过程文件

    HEADER_ROW = 1
    DATA_START_ROW = 2

    def __init__(self, file_path: str, sheet_name: str = "工单列表"):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_tickets(self) -> List[Dict[str, Any]]:
        """从Excel读取工单列表

        Returns:
            工单列表 [{"ticket_id": "xxx", "note": "yyy", "row": 2}, ...]
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
            return []

        if not os.path.exists(self.file_path):
            logger.error(f"❌ Excel文件不存在: {self.file_path}")
            return []

        wb = load_workbook(self.file_path)
        if self.sheet_name not in wb.sheetnames:
            # 使用第一个sheet
            ws = wb.active
            logger.warning(f"⚠️ 未找到Sheet '{self.sheet_name}'，使用: {ws.title}")
        else:
            ws = wb[self.sheet_name]

        tickets = []
        for row in range(self.DATA_START_ROW, ws.max_row + 1):
            ticket_id = ws.cell(row=row, column=self.COL_TICKET).value
            if not ticket_id:
                continue

            ticket_id = str(ticket_id).strip()
            note = ws.cell(row=row, column=self.COL_NOTE).value or ""
            status = ws.cell(row=row, column=self.COL_STATUS).value or ""

            # 不跳过已完成工单——用户可能需要重复执行
            if status in ("成功", "完成"):
                logger.debug(f"工单 {ticket_id} 上次状态: {status}，本次重新执行")

            tickets.append({
                "ticket_id": ticket_id,
                "note": str(note).strip(),
                "row": row,
            })

        wb.close()
        logger.info(f"📊 从Excel读取 {len(tickets)} 个待处理工单")
        return tickets

    def update_result(self, row: int, status: str, result_text: str,
                      time_text: str, process_file: str):
        """回写结果到Excel"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.file_path)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.active

            ws.cell(row=row, column=self.COL_STATUS, value=status)
            ws.cell(row=row, column=self.COL_RESULT, value=result_text)
            ws.cell(row=row, column=self.COL_TIME, value=time_text)
            ws.cell(row=row, column=self.COL_PROCESS, value=process_file)

            wb.save(self.file_path)
            wb.close()
            logger.info(f"📝 Excel已更新: 行{row} → {status}")
        except Exception as e:
            logger.error(f"❌ Excel回写失败: {e}")

    @staticmethod
    def create_template(file_path: str):
        """创建Excel模板"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "工单列表"

            # 表头
            headers = ["工单号", "补充说明", "状态", "执行结果", "执行时间", "过程文件"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 列宽
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 50

            # 示例数据
            ws.cell(row=2, column=1, value="ONES-12345")
            ws.cell(row=2, column=2, value="修复登录页面样式问题")

            wb.save(file_path)
            wb.close()
            logger.info(f"✅ Excel模板已创建: {file_path}")
        except ImportError:
            logger.error("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")


# ============================================================
#  过程日志记录器
# ============================================================
class ProcessLogger:
    """过程日志管理器"""

    def __init__(self, process_dir: str = "process"):
        self.process_dir = process_dir
        try:
            os.makedirs(self.process_dir, exist_ok=True)
        except PermissionError:
            # 无写权限（如 /opt/lango/ 下运行），回退到用户目录
            fallback = os.path.join(os.path.expanduser("~"), "ones-ai-process")
            logger.warning(f"⚠️ 无法创建 {self.process_dir}，回退到 {fallback}")
            self.process_dir = fallback
            os.makedirs(self.process_dir, exist_ok=True)

    def save(self, task_id: str, task: Dict[str, Any],
             result: Dict[str, Any]) -> str:
        """保存过程日志

        Returns:
            日志文件路径
        """
        success = result.get('success', False)
        status_text = "成功" if success else "失败"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{task_id}_{status_text}_{timestamp}.md"
        filepath = os.path.join(self.process_dir, filename)

        duration = result.get('duration', 0)

        # 提取开发者补充说明（如果在prompts中有标记）
        prompts_text = task.get('prompts', 'N/A')
        dev_note_section = ""
        if '【开发者补充】' in str(prompts_text):
            dev_note_section = f"""\n## 开发者补充说明\n\n{prompts_text.split('【开发者补充】')[-1].strip()}\n"""

        content = f"""# 工单处理报告: {task_id}

## 基本信息

| 项目 | 内容 |
|------|------|
| 工单号 | {task_id} |
| 任务名称 | {task.get('taskName', 'N/A')} |
| 项目名称 | {task.get('projectName', 'N/A')} |
| 执行状态 | {'✅ 成功' if success else '❌ 失败'} |
| 执行耗时 | {duration:.1f} 秒 |
| 执行时间 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} |
| 工作目录 | {os.getcwd()} |
{dev_note_section}
## 任务描述

{task.get('description', 'N/A')}

## 关键词

{task.get('keywords', 'N/A')}

## 提示词

{prompts_text}

## 执行输出

```
{result.get('stdout', '')[:5000]}
```

"""

        if result.get('stderr'):
            content += f"""## 错误输出

```
{result.get('stderr', '')[:2000]}
```

"""

        if result.get('error'):
            content += f"""## 异常信息

- 错误类型: {result.get('exception_type', 'Unknown')}
- 错误详情: {result.get('error', '')}

"""

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)

        logger.info(f"📄 过程日志已保存: {filepath}")
        return filepath


# ============================================================
#  使用追踪器
# ============================================================
class UsageTracker:
    """使用记录追踪器"""

    def __init__(self, config: Config):
        self.enabled = config.get('tracking.enabled', True)
        self.api_url = config.get('tracking.api_url', '')
        self.local_log = config.get('tracking.local_log', 'usage_log.json')
        self._records = []

    def _get_username(self):
        """获取当前用户名（兼容 Docker 容器、Linux 宿主机、Windows）"""
        # 优先级: USER -> USERNAME -> whoami -> getpass
        user = os.environ.get('USER') or os.environ.get('USERNAME')
        if user and user != 'root':
            return user
        # 容器内可能是 root，尝试 whoami
        try:
            import subprocess
            user = subprocess.check_output(['whoami'], timeout=3).decode().strip()
            if user and user != 'root':
                return user
        except Exception:
            pass
        # 最后兜底
        try:
            import getpass
            return getpass.getuser()
        except Exception:
            return user or 'unknown'

    def _get_hostname(self):
        """获取主机名"""
        hostname = os.environ.get('HOSTNAME') or os.environ.get('COMPUTERNAME')
        if hostname:
            return hostname
        try:
            import socket
            return socket.gethostname()
        except Exception:
            return 'unknown'

    def record_invocation(self, tickets: List[str], mode: str):
        """记录一次调用"""
        if not self.enabled:
            return

        record = {
            "user": self._get_username(),
            "machine": self._get_hostname(),
            "invoke_time": datetime.now().isoformat(),
            "mode": mode,
            "tickets": tickets,
            "cwd": os.getcwd(),
            "ticket_details": [],
        }
        self._records.append(record)

    def record_result(self, results_summary: Dict[str, int], duration: float):
        """记录执行结果"""
        if not self.enabled or not self._records:
            return

        self._records[-1].update({
            "results_summary": results_summary,
            "duration_seconds": round(duration, 1),
            "finish_time": datetime.now().isoformat(),
        })

    def record_ticket_detail(self, ticket_id: str, note: str,
                              success: bool, duration: float):
        """记录单个工单的执行详情"""
        if not self.enabled or not self._records:
            return
        detail = {
            "ticket_id": ticket_id,
            "note": note,
            "status": "成功" if success else "失败",
            "duration": round(duration, 1),
        }
        self._records[-1].setdefault("ticket_details", []).append(detail)

    def flush(self):
        """提交追踪记录"""
        if not self.enabled or not self._records:
            return

        # 写本地文件
        self._write_local()

        # 上报API（如配置了）
        if self.api_url:
            self._send_to_api()

    def _write_local(self):
        """写入本地JSON文件

        注意：未加文件锁，多个 ones-AI 实例同时运行时可能互相覆盖。
        当前场景（单用户单实例）下不会出问题。
        """
        try:
            existing = []
            if os.path.exists(self.local_log):
                with open(self.local_log, 'r', encoding='utf-8') as f:
                    existing = json.load(f)

            existing.extend(self._records)

            with open(self.local_log, 'w', encoding='utf-8') as f:
                json.dump(existing, f, ensure_ascii=False, indent=2)

            logger.info(f"📊 使用记录已保存: {self.local_log}")
        except Exception as e:
            logger.error(f"❌ 使用记录写入失败: {e}")

    def _send_to_api(self):
        """上报到追踪API"""
        try:
            for record in self._records:
                response = requests.post(
                    self.api_url,
                    json=record,
                    timeout=5,
                )
                if response.status_code == 200:
                    logger.info(f"📡 使用记录已上报API")
                else:
                    logger.warning(f"⚠️ API上报异常: HTTP {response.status_code}")
        except Exception as e:
            logger.warning(f"⚠️ API上报失败（不影响执行）: {e}")


# ============================================================
#  主流程编排
# ============================================================
class TaskRunner:
    """工单任务执行编排器"""

    def __init__(self, config: Config):
        self.config = config
        self.executor = ClaudeExecutor(config)
        self.process_logger = ProcessLogger(config.get('output.process_dir', 'process'))
        self.tracker = UsageTracker(config)
        self.excel_manager = None

    def run(self, tickets: List[str], excel_path: str = "",
             cli_notes: Dict[str, str] = None, cli_code_dirs: Dict[str, str] = None,
             json_output: bool = False):
        """主执行流程

        Args:
            tickets: 工单号列表
            excel_path: Excel文件路径（可选）
            cli_notes: 命令行传入的补充说明（可选）
            cli_code_dirs: 命令行传入的代码位置（可选）
            json_output: 是否输出JSON格式结果（供远程调用解析）
        """
        start_time = time.time()
        results_summary = {"total": 0, "success": 0, "failed": 0}

        # Excel管理器
        excel_rows = {}
        excel_notes = cli_notes or {}  # 优先使用命令行notes，Excel模式会覆盖
        code_dirs = cli_code_dirs or {}
        if excel_path:
            sheet_name = self.config.get('excel.sheet_name', '工单列表')
            self.excel_manager = ExcelManager(excel_path, sheet_name)

            if not tickets:
                # 从Excel读取工单号
                excel_tickets = self.excel_manager.read_tickets()
                tickets = [t["ticket_id"] for t in excel_tickets]
                excel_rows = {t["ticket_id"]: t["row"] for t in excel_tickets}
                excel_notes = {t["ticket_id"]: t["note"] for t in excel_tickets if t["note"]}

        if not tickets:
            logger.warning("⚠️ 没有待处理的工单")
            return

        ui()
        header_line(f"任务队列 · 共 {len(tickets)} 个工单")
        for i, t in enumerate(tickets, 1):
            ui(f"  {i:>2}.  {t}")
        header_line()
        ui()

        # 记录使用情况
        self.tracker.record_invocation(tickets, 'local')

        # 逐个处理工单
        for idx, ticket_id in enumerate(tickets, 1):
            ui()
            header_line(f"[{idx}/{len(tickets)}]  {ticket_id}")

            results_summary["total"] += 1
            task_start = datetime.now()

            # 更新Excel状态
            if self.excel_manager and ticket_id in excel_rows:
                self.excel_manager.update_result(
                    excel_rows[ticket_id], "执行中", "", "", ""
                )

            # 1. 构造任务（票号+代码位置+补充说明）
            dev_note = excel_notes.get(ticket_id, '')
            code_dir = code_dirs.get(ticket_id, '')

            # 构建提示词（含代码位置指引）
            prompt_parts = [f"请处理ONES工单 {ticket_id}。"]
            if code_dir:
                prompt_parts.append(f"【代码位置】项目代码在 {code_dir} 目录，请在该目录下进行分析和修复。")
            if dev_note:
                prompt_parts.append(f"【开发者补充】{dev_note}")
            if not code_dir and not dev_note:
                prompt_parts.append("请根据工单内容自行分析和处理。")

            task = {
                "taskId": ticket_id,
                "taskName": ticket_id,
                "projectName": "N/A",
                "codeDirectory": code_dir,
                "prompts": ''.join(prompt_parts),
            }

            # 2. 执行任务
            result = self.executor.execute(task)

            # 3. 保存过程日志
            process_file = self.process_logger.save(ticket_id, task, result)

            # 4. 计数
            if result.get('success'):
                results_summary["success"] += 1
            else:
                results_summary["failed"] += 1

            # 4.5 记录工单详情到追踪器
            self.tracker.record_ticket_detail(
                ticket_id=ticket_id,
                note=dev_note,
                success=result.get('success', False),
                duration=result.get('duration', 0),
            )

            # 4.6 JSON 输出（供远程调用方解析）
            if json_output:
                details = self._extract_result_details(result, ticket_id)
                json_line = json.dumps({
                    "ticket_id": ticket_id,
                    "status": "success" if result.get('success') else "failed",
                    "duration": round(result.get('duration', 0), 1),
                    "summary": details.get('summary', ''),
                    "title": details.get('title', ''),
                    "conclusion": details.get('conclusion', ''),
                    "analysis": details.get('analysis', ''),
                    "report_path": details.get('report_path', ''),
                }, ensure_ascii=False)
                print(json_line, flush=True)

            # 5. 回写Excel
            if self.excel_manager and ticket_id in excel_rows:
                status = "成功" if result['success'] else "失败"
                result_text = self._build_result_text(result)
                time_text = f"{task_start.strftime('%H:%M')} → {datetime.now().strftime('%H:%M')}"
                self.excel_manager.update_result(
                    excel_rows[ticket_id], status, result_text,
                    time_text, process_file
                )

            # 工单间间隔（防止过快）
            if idx < len(tickets):
                time.sleep(2)

        # 汇总报告
        total_duration = time.time() - start_time
        self.tracker.record_result(results_summary, total_duration)
        self.tracker.flush()

        self._print_summary(results_summary, total_duration)



    def _write_failure(self, ticket_id: str, excel_rows: dict,
                       task_start: datetime, error_msg: str):
        """写入失败记录"""
        fake_task = {"taskId": ticket_id, "taskName": "N/A"}
        fake_result = {"success": False, "error": error_msg, "duration": 0}
        process_file = self.process_logger.save(ticket_id, fake_task, fake_result)

        if self.excel_manager and ticket_id in excel_rows:
            time_text = f"{task_start.strftime('%H:%M')} → {datetime.now().strftime('%H:%M')}"
            self.excel_manager.update_result(
                excel_rows[ticket_id], "失败", error_msg,
                time_text, process_file
            )

    @staticmethod
    def _build_result_text(result: Dict[str, Any]) -> str:
        """构建结果概要文本"""
        if result.get('success'):
            duration = result.get('duration', 0)
            return f"执行成功 (耗时 {duration:.0f}s)"
        else:
            error = result.get('error') or result.get('stderr') or '未知错误'
            if len(str(error)) > 100:
                error = str(error)[:100] + "..."
            return f"执行失败: {error}"

    def _extract_result_details(self, result: Dict[str, Any],
                                 ticket_id: str) -> Dict[str, str]:
        """从 claude 执行结果中提取结构化详情

        提取字段：
          - title: 工单标题（从输出中匹配）
          - conclusion: 执行结论摘要
          - analysis: 完整分析内容（截断到合理长度）
          - report_path: 报告文件路径
          - summary: 综合摘要
        """
        details: Dict[str, str] = {
            'title': '',
            'conclusion': '',
            'analysis': '',
            'report_path': '',
            'summary': self._build_result_text(result),
        }

        stdout = result.get('stdout', '')
        if not stdout:
            return details

        # 提取标题（匹配 **标题** / 标题: / # 任务报告: 等模式）
        title_patterns = [
            r'\*\*标题\*\*[\s:：]*(.+)',
            r'[#]+\s*任务报告[：:]\s*(.+)',
            r'\|\s*\*\*标题\*\*\s*\|\s*(.+?)\s*\|',
            r'\|\s*标题\s*\|\s*(.+?)\s*\|',
            r'\*\*工单号\*\*.*?\|.*?\*\*(.+?)\*\*',
        ]
        for pat in title_patterns:
            m = re.search(pat, stdout)
            if m:
                details['title'] = m.group(1).strip()[:100]
                break

        # 提取结论/执行摘要（匹配 ## 执行摘要 / ## ✅ 任务完成 等）
        conclusion_patterns = [
            r'##\s*(?:✅\s*)?任务完成[\s\S]*?\n\n(.+?)(?:\n\n|\n##|$)',
            r'##\s*执行摘要\s*\n+(.+?)(?:\n\n|\n##|$)',
            r'\*\*状态\*\*[\s:：]*(已完成|完成|部分完成|无法完成)',
        ]
        for pat in conclusion_patterns:
            m = re.search(pat, stdout, re.DOTALL)
            if m:
                conclusion = m.group(1).strip()
                # 清理 markdown 格式
                conclusion = re.sub(r'\*\*(.+?)\*\*', r'\1', conclusion)
                conclusion = re.sub(r'\|', ' ', conclusion)
                details['conclusion'] = conclusion[:200]
                break

        # 提取分析内容（取 stdout 的核心部分，去掉过多前后缀）
        analysis = stdout.strip()
        # 去掉 markdown 分隔线和头部装饰
        analysis = re.sub(r'^[-─=]{3,}.*$', '', analysis, flags=re.MULTILINE)
        analysis = analysis.strip()
        if len(analysis) > 2000:
            analysis = analysis[:2000] + '\n... (输出已截断)'
        details['analysis'] = analysis

        # 检测报告文件路径
        report_patterns = [
            r'报告位置[\s:：]*[`\n]*([^\s`]+\.md)',
            r'workspace/doc/[^\s]*?/report/\d+\.md',
            r'/home/[^\s]*?/workspace/doc/[^\s]*?/report/\d+\.md',
        ]
        for pat in report_patterns:
            m = re.search(pat, stdout)
            if m:
                path = m.group(0) if '/' in m.group(0) else m.group(1)
                details['report_path'] = path.strip()
                break

        # 如果没从输出中找到报告路径，检查默认位置
        if not details['report_path']:
            default_report = os.path.join(
                os.getcwd(), 'workspace', 'doc', ticket_id, 'report', '1.md'
            )
            if os.path.exists(default_report):
                details['report_path'] = default_report

        # 改善 summary：如果成功且有 conclusion，用 conclusion 替代默认摘要
        if result.get('success') and details['conclusion']:
            duration = result.get('duration', 0)
            details['summary'] = f"{details['conclusion']} (耗时 {duration:.0f}s)"

        return details

    @staticmethod
    def _print_summary(summary: Dict[str, int], duration: float):
        """打印执行汇总"""
        total = summary['total']
        ok = summary['success']
        fail = summary['failed']
        mins = duration / 60

        ui()
        header_line("执行汇总")
        ui(f"  工单总数    {total}")
        ui(f"  成功        {ok}  ✅")
        ui(f"  失败        {fail}  {'❌' if fail else ''}")
        ui(f"  总耗时      {duration:.0f}s ({mins:.1f} 分钟)")
        header_line()
        if fail == 0:
            ui("\n🎉 全部工单处理完毕，无失败。")
        else:
            ui(f"\n⚠️  {fail} 个工单处理失败，请检查 process/ 目录的过程日志。")
        ui()


# ============================================================
#  命令行入口
# ============================================================
def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description="ONES工单夜间自动处理工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 命令行指定工单号
  python ones_task_runner.py --tickets ONES-12345 ONES-12346

  # 带补充说明
  python ones_task_runner.py --tickets ONES-12345 --notes "检查Captcha组件"

  # 从Excel读取
  python ones_task_runner.py --excel tasks.xlsx

  # 交互式输入
  python ones_task_runner.py

  # 生成Excel模板
  python ones_task_runner.py --create-template
        """
    )

    parser.add_argument(
        '--tickets', '-t', nargs='+',
        help='工单号列表（空格分隔）'
    )
    parser.add_argument(
        '--notes', '-n', nargs='+',
        help='补充说明（与--tickets一一对应，空格分隔）'
    )
    parser.add_argument(
        '--code-dirs', '-d', nargs='+',
        help='代码位置（与--tickets一一对应，每个工单对应的代码仓库路径）'
    )
    parser.add_argument(
        '--json-output', action='store_true',
        help='以 JSON 格式输出每个工单的执行结果（供远程调用方解析）'
    )
    parser.add_argument(
        '--excel', '-e',
        help='Excel文件路径'
    )
    parser.add_argument(
        '--config', '-c', default='config.yaml',
        help='配置文件路径（默认: config.yaml）'
    )
    parser.add_argument(
        '--create-template',
        action='store_true',
        help='创建Excel模板文件'
    )
    parser.add_argument(
        '--agent-dir',
        help='Agent Teams 工作目录（由 ones-AI 脚本传入，用于 Claude subagent 资源定位）'
    )
    parser.add_argument(
        '--task-mode',
        default='fix',
        help='任务模式: fix(修复), analyze(分析), auto(自动) 等'
    )
    parser.add_argument(
        '--extra-mounts',
        help='额外挂载路径（逗号分隔，由 ones-AI 脚本处理 Docker 挂载）'
    )

    return parser.parse_args()


def interactive_input() -> List[str]:
    """交互式输入工单号"""
    ui()
    header_line("交互模式")
    ui("  输入工单号（逗号/空格分隔，q 退出）")
    ui()

    raw = input("> ").strip()
    if raw.lower() in ('q', 'quit', 'exit'):
        ui("已退出。")
        sys.exit(0)

    # 支持逗号、空格、分号分隔，纯数字自动补 ONES-
    raw_parts = [t.strip() for t in re.split(r'[,;\s]+', raw) if t.strip()]
    tickets = [normalize_ticket_id(t) for t in raw_parts]

    if not tickets:
        ui("❌ 未输入任何工单号")
        sys.exit(1)

    return tickets


def main():
    """主入口"""
    args = parse_args()

    # 加载配置
    config = Config(args.config)

    # 设置日志级别
    log_level = config.get('output.log_level', 'INFO')
    logger.setLevel(getattr(logging, log_level, logging.INFO))

    # 创建模板
    if args.create_template:
        template_path = config.get('excel.default_file', 'tasks.xlsx')
        ExcelManager.create_template(template_path)
        return

    ui()
    ui("╔══════════════════════════════════════════════════════╗")
    ui("║            ones-AI · ONES 工单自动处理               ║")
    ui("╚══════════════════════════════════════════════════════╝")
    ui(f"  版本   1.0.0")
    ui(f"  目录   {os.getcwd()}")
    ui(f"  时间   {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # 确定工单来源
    tickets = []
    excel_path = ""
    cli_notes = {}  # 命令行模式的补充说明

    if args.tickets:
        # 命令行指定（支持逗号分隔、纯数字自动补 ONES-）
        tickets = normalize_ticket_list(args.tickets)
        # 处理命令行补充说明（--notes）
        if args.notes:
            for i, note in enumerate(args.notes):
                if i < len(tickets):
                    cli_notes[tickets[i]] = note
        # 处理命令行代码位置（--code-dirs）
        cli_code_dirs = {}
        if args.code_dirs:
            for i, d in enumerate(args.code_dirs):
                if i < len(tickets):
                    cli_code_dirs[tickets[i]] = d
    elif args.excel:
        # Excel文件
        excel_path = args.excel
    else:
        # 交互式输入
        tickets = interactive_input()

    # 执行
    runner = TaskRunner(config)
    runner.run(tickets, excel_path, cli_notes,
               cli_code_dirs if 'cli_code_dirs' in dir() else None,
               json_output=getattr(args, 'json_output', False))


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        logger.warning("\n\n⚠️ 用户中断，正在退出...")
        sys.exit(1)
    except Exception as e:
        logger.error(f"\n\n❌ 未预期错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

